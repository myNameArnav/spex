from requests.auth import HTTPBasicAuth
from requests_oauthlib import OAuth2Session
from openpyxl import load_workbook

client_id = 'd3a47d91f6ea425192e833b4e6c5e688'
client_secret = '0f27fcf6e9d54260bf92f64f2da8537d'
redirect_uri = 'https://developer.spotify.com/'
user_id = '31zc6ritsk5fyws7f4vkhnekloxq'

authorization_base_url = "https://accounts.spotify.com/authorize"
token_url = "https://accounts.spotify.com/api/token"
scope = [
   "user-read-email",
   "playlist-read-collaborative",
   "playlist-read-private"
]

spotify = OAuth2Session(client_id, scope=scope, redirect_uri=redirect_uri)

authorization_url, state = spotify.authorization_url(authorization_base_url)
print('Please go here and authorize: ', authorization_url)

redirect_response = input('\n\nPaste the full redirect URL here: ')

auth = HTTPBasicAuth(client_id, client_secret)

token = spotify.fetch_token(token_url, auth=auth, authorization_response=redirect_response)
access_token = token["access_token"]

user = 'https://api.spotify.com/v1/users/' + user_id + '/playlists'
r = spotify.get(user)

playlists = r.json()

link = "Spotify.xlsx"
wb = load_workbook(filename=link)

index = 1

for i in range(len(playlists["items"])):
    
    playlist_link = playlists["items"][i]['href']
    
    r = spotify.get(playlist_link)

    playlist_details = r.json()
    playlist_name = playlist_details["name"]
    
    endpoint = playlist_link + '/tracks'
    r = spotify.get(endpoint)
    tracks = r.json()
    
    if playlist_name in wb.worksheets:
        sheet = wb[playlist_name]
        #print(sheet)
    else:
        wb.create_sheet(playlist_name)
        wb.save(link)
        wb = load_workbook(filename=link)
        sheet = wb[playlist_name]
        #print(sheet)
    
    sheetAll = wb["ALL"]
    
    for i in range(len(tracks["items"])):
        track_link = tracks["items"][i]["track"]["album"]["artists"][0]["external_urls"]["spotify"]
        
        c1 = sheetAll.cell(row=index, column=1)
        c1.value = track_link
        
        c2 = sheet.cell(row=(i+1), column=1)
        c2.value = track_link
        
        #print(track_link)
        
        index += 1
    
    #print(playlist_link)
    
wb.save(link)
print("Done")